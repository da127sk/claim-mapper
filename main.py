import streamlit as st
import pandas as pd
import io
from openpyxl.styles import numbers
from openpyxl import load_workbook

# Internal standard names and their descriptive headers
standard_columns = {
    'Policy Year': 'Policy Year',
    'Status': 'Claim Status',
    'Body Part Category': 'Body Part Cat.',
    'Injury Cause Category': 'Injury Cause Cat.',
    'Incurred': 'Incurred',
    'Litigation Status': 'Litigation'
}

required_fields = ['Policy Year', 'Status', 'Body Part Category', 'Injury Cause Category', 'Incurred']
optional_fields_with_default = {'Litigation Status': 'N/A'}

st.set_page_config(layout="wide")
st.title("Claim Mapper")

st.markdown(
    """
    <div style='
        background-color: #fff3cd;
        border-left: 6px solid #ffc107;
        padding: 10px 16px;
        margin-bottom: 20px;
        font-size: 16px;
        color: #856404;
        font-weight: bold;
        border-radius: 4px;'
    >
        ⚠️ <strong>Disclaimer:</strong> Before uploading, please make sure the loss run has at least <u>4 years</u> worth of losses counting from the newest year.
    </div>
    """,
    unsafe_allow_html=True
)

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    st.session_state['file_data'] = uploaded_file

if 'file_data' in st.session_state:
    xlsx = pd.ExcelFile(st.session_state['file_data'])
    sheet_name = st.selectbox("Select a worksheet to load", xlsx.sheet_names)
    df = pd.read_excel(xlsx, sheet_name=sheet_name)
    df.columns = df.columns.str.strip()
    st.write("Preview of uploaded data:")
    st.dataframe(df.head())

    st.subheader("Step 1: Map Columns")

    if st.button("🔄 Restart Mapping"):
        st.session_state.pop('mapping', None)

    if 'mapping' not in st.session_state:
        st.session_state['mapping'] = {}

    mapping = st.session_state['mapping']

    for internal_name, description in standard_columns.items():
        default = mapping.get(internal_name)
        options = [None] + list(df.columns)
        selected = st.selectbox(
            f"Select column for '{internal_name}' ({description})",
            options,
            index=(options.index(default) if default in options else 0),
            key=f"map_{internal_name}"
        )
        if selected:
            mapping[internal_name] = selected
        elif internal_name in mapping:
            del mapping[internal_name]

    file_name = st.text_input("Enter export file name:", value="claim_output.xlsx")
    if not file_name.lower().endswith(".xlsx"):
        file_name += ".xlsx"

    missing_required = [col for col in required_fields if col not in mapping]
    if missing_required:
        st.error(f"Missing required column mappings: {', '.join(missing_required)}")
    else:
        st.subheader("Step 2: Enter MOD values and Payrolls for 4 years + Projection")
        mod_data = []
        labels = ["Projection Year", "Current Year", "Policy Year", "Policy Year", "Policy Year"]

        col1, col2, col3 = st.columns([1, 1, 1])
        projection_year_str = col1.text_input(label="Projection Year", placeholder="e.g. 2025", key="mod_year_0")
        mod = col2.text_input(label="Mod", placeholder="Mod", key="mod_value_0")
        col3.markdown("0 (locked)")
        payroll = "0"

        try:
            projection_year = int(projection_year_str)
        except:
            projection_year = None

        if projection_year_str and mod:
            mod_data.append({
                "Policy Year": projection_year_str,
                "Mod": mod,
                "Total Payroll": payroll
            })

        for i in range(1, 5):
            col1, col2, col3 = st.columns([1, 1, 1])

            year_label = labels[i]
            suggested_year = str(projection_year - i) if projection_year is not None else ""

            year = col1.text_input(label=year_label, value=suggested_year, placeholder="Policy Year", key=f"mod_year_{i}")
            mod = col2.text_input(label="Mod", placeholder="Mod", key=f"mod_value_{i}")
            payroll = col3.text_input(label="Payroll", placeholder="Payroll", key=f"payroll_{i}")

            if year and mod:
                mod_data.append({
                    "Policy Year": year,
                    "Mod": mod,
                    "Total Payroll": payroll
                })




        mod_df_raw = pd.DataFrame(mod_data)

        if st.button("Export Mapped Columns"):
            used_columns = [mapping[k] for k in mapping]
            new_df = df[used_columns].copy()
            new_df.columns = [standard_columns[k] for k in mapping.keys()]

            for field, default_val in optional_fields_with_default.items():
                if field not in mapping:
                    new_df[standard_columns[field]] = default_val
                    st.warning("Litigation status not provided. This will not be reflected in the dashboard.")

            if "Policy Year" in new_df.columns:
                col = new_df["Policy Year"]
                if pd.api.types.is_datetime64_any_dtype(col):
                    new_df["Policy Year"] = col.dt.year
                elif pd.api.types.is_numeric_dtype(col):
                    new_df["Policy Year"] = col.astype(int)
                else:
                    parsed = pd.to_datetime(col, errors="coerce")
                    new_df["Policy Year"] = parsed.dt.year.fillna(col).astype(int)

            #### 🧮 Normalize Status
            def normalize_status(status):
                status = str(status).strip().lower()
                if status in ['Opened', 'reopen', 're-open', 're opened', 'reopened']:
                    return 'Open'
                elif status in ['Close', 'reclosed', 're-closed', 're closed']:
                    return 'Closed'
                elif 'open' in status:
                    return 'Open'
                elif 'closed' in status:
                    return 'Closed'
                return 'Unknown'

            temp_df = new_df.copy()
            if "Claim Status" in temp_df.columns:
                temp_df['Normalized Status'] = temp_df['Claim Status'].apply(normalize_status)

            if "Policy Year" in temp_df.columns and "Normalized Status" in temp_df.columns:
                status_summary = (
                    temp_df.groupby(['Policy Year', 'Normalized Status'])
                        .size()
                        .reset_index(name='Value')
                        .rename(columns={'Normalized Status': 'Claim Type'})
                )
            else:
                status_summary = pd.DataFrame()

            if "Policy Year" in temp_df.columns and "Incurred" in temp_df.columns:
                incurred_summary = (
                    temp_df.groupby('Policy Year')['Incurred']
                        .sum(numeric_only=True)
                        .reset_index(name='Value')
                )
                incurred_summary['Claim Type'] = 'Total Incurred'
            else:
                incurred_summary = pd.DataFrame()

            summary = pd.concat([status_summary, incurred_summary], ignore_index=True)

            ### 🧮 Format Payroll Input
            mod_df_raw["Policy Year"] = mod_df_raw["Policy Year"].astype(int)
            mod_df_raw["Mod"] = mod_df_raw["Mod"].astype(float).round(0)
            mod_df_raw["Total Payroll"] = mod_df_raw["Total Payroll"].astype(float)

            sorted_mod = mod_df_raw.sort_values("Policy Year", ascending=False).reset_index(drop=True)
            mod_history = sorted_mod[["Policy Year", "Mod"]].copy()
            payroll_df = sorted_mod[["Policy Year", "Total Payroll"]].copy()

            labels = ["Projected", "Current"] + sorted_mod["Policy Year"].iloc[2:].astype(str).tolist()
            mod_history["Policy Year"] = labels

            # ⬇ Use only 4 years in summary and new_df
            historical_years = sorted_mod["Policy Year"].iloc[1:].tolist()
            new_df = new_df[new_df["Policy Year"].isin(historical_years)]
            summary = summary[summary["Policy Year"].isin(historical_years)]
            summary = summary.pivot(index="Policy Year", columns="Claim Type", values="Value").reset_index()
            summary["Total Claims"] = summary.get("Open", 0).fillna(0) + summary.get("Closed", 0).fillna(0)

            summary = summary.rename(columns={"Incurred": "Total Incurred"})
            summary = summary.merge(payroll_df, on="Policy Year", how="left")
            summary = summary.rename(columns={"Total Payroll": "Total Payroll"})
            summary = summary[["Policy Year", "Closed", "Open", "Total Claims", "Total Incurred", "Total Payroll"]]
            summary = summary.sort_values("Policy Year", ascending=False)

            st.subheader("📄 Data Preview")
            st.dataframe(new_df.head())

            st.subheader("📄 Claim Summary Preview")
            st.dataframe(summary)

            st.subheader("📄 Mod History Preview")
            st.dataframe(mod_history)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                new_df.to_excel(writer, index=False, sheet_name="Data")
                summary.to_excel(writer, index=False, sheet_name="OpenClosedIncurred")
                mod_history.to_excel(writer, index=False, sheet_name="Mod History")

            output.seek(0)
            wb = load_workbook(output)

            # 💵 Format Incurred + Payroll in OpenClosedIncurred
            ws = wb["OpenClosedIncurred"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for col_name in ["Total Incurred", "Total Payroll"]:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if col_name == "Total Payroll":
                                cell.number_format = '"$"#,##0'
                            else:
                                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            # Format Mod History worksheet
            ws = wb["Mod History"]
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = '0'

            final_output = io.BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            st.download_button("Download Excel File", final_output, file_name=file_name)
